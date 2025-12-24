"""
Excel生成工具
用于将测试用例生成结果导出为Excel文件
"""
import json
import os
import urllib.parse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from langchain_core.tools import tool
from source.agent.utils.log_utils import MyLogger

log = MyLogger().get_logger()


def _format_test_steps(steps: Any) -> str:
    """格式化测试步骤"""
    if isinstance(steps, list):
        return "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps)])
    elif isinstance(steps, str):
        return steps
    else:
        return str(steps)


@tool
def generate_excel_from_test_cases(
    test_cases_json: str,
    output_path: Optional[str] = None,
    review_result_json: Optional[str] = None
) -> str:
    """从测试用例JSON生成Excel文件
    
    Args:
        test_cases_json: JSON格式的测试用例列表
        output_path: 输出Excel文件路径（可选，默认生成在项目根目录）
        review_result_json: JSON格式的评审结果（可选）
    
    Returns:
        生成结果信息
    """
    try:
        # 解析测试用例
        test_cases = json.loads(test_cases_json)
        if isinstance(test_cases, dict):
            test_cases = [test_cases]
        elif not isinstance(test_cases, list):
            test_cases = [test_cases]
        
        if not test_cases:
            return "错误：测试用例列表为空"
        
        # 解析评审结果（如果有）
        review_result = None
        if review_result_json:
            try:
                review_result = json.loads(review_result_json)
            except:
                pass
        
        # 确定输出路径
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            project_root = Path(__file__).parent.parent.parent.parent
            downloads_dir = project_root / "downloads"
            downloads_dir.mkdir(parents=True, exist_ok=True)
            filename = f"测试用例_{timestamp}.xlsx"
            output_path = str(downloads_dir / filename)
        else:
            filename = Path(output_path).name
        
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 设置表头
        headers = [
            "用例ID",
            "测试类型",
            "用例描述",
            "前置条件",
            "测试步骤",
            "预期结果",
            "优先级"
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入测试用例数据
        for row_idx, test_case in enumerate(test_cases, start=2):
            ws.cell(row=row_idx, column=1, value=test_case.get("test_case_id", ""))
            ws.cell(row=row_idx, column=2, value=test_case.get("test_type", ""))
            ws.cell(row=row_idx, column=3, value=test_case.get("test_description", ""))
            ws.cell(row=row_idx, column=4, value=test_case.get("preconditions", ""))
            ws.cell(row=row_idx, column=5, value=_format_test_steps(test_case.get("test_steps", [])))
            ws.cell(row=row_idx, column=6, value=test_case.get("expected_result", ""))
            ws.cell(row=row_idx, column=7, value=test_case.get("priority", ""))
            
            # 设置单元格样式
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        
        # 设置列宽
        column_widths = {
            "A": 12,  # 用例ID
            "B": 12,  # 测试类型
            "C": 30,  # 用例描述
            "D": 20,  # 前置条件
            "E": 40,  # 测试步骤
            "F": 30,  # 预期结果
            "G": 10,  # 优先级
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 设置行高
        ws.row_dimensions[1].height = 30  # 表头行高
        for row_idx in range(2, len(test_cases) + 2):
            ws.row_dimensions[row_idx].height = 60  # 数据行高
        
        # 如果有评审结果，添加评审信息工作表
        if review_result:
            review_ws = wb.create_sheet("评审结果")
            review_headers = ["评审项", "得分", "说明"]
            review_data = [
                ["覆盖率", review_result.get("coverage_score", 0), ""],
                ["可执行性", review_result.get("executability_score", 0), ""],
                ["无歧义性", review_result.get("clarity_score", 0), ""],
                ["总分", review_result.get("score", 0), ""],
                ["是否通过", "是" if review_result.get("is_passed", False) else "否", ""],
            ]
            
            # 写入评审结果表头
            for col_idx, header in enumerate(review_headers, start=1):
                cell = review_ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # 写入评审结果数据
            for row_idx, row_data in enumerate(review_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = review_ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
            
            # 如果有优化建议，添加到评审结果中
            suggestions = review_result.get("suggestions", [])
            if suggestions:
                review_ws.cell(row=7, column=1, value="优化建议").font = header_font
                for idx, suggestion in enumerate(suggestions, start=1):
                    review_ws.cell(row=7 + idx, column=1, value=f"{idx}. {suggestion}")
                    review_ws.cell(row=7 + idx, column=1).alignment = Alignment(wrap_text=True)
            
            # 设置评审结果工作表列宽
            review_ws.column_dimensions["A"].width = 15
            review_ws.column_dimensions["B"].width = 10
            review_ws.column_dimensions["C"].width = 50
        
        # 保存Excel文件
        wb.save(output_path)
        
        # 生成下载链接
        # 注意：链接应该指向 FastAPI 后端（端口 9501），而不是 Streamlit（端口 8501）
        filename = Path(output_path).name
        # URL 编码文件名以支持中文
        encoded_filename = urllib.parse.quote(filename, safe='')
        # 使用完整的后端 API 地址
        import os
        api_base_url = os.getenv("API_BASE_URL", "http://localhost:9501")
        download_url = f"{api_base_url}/api/download/{encoded_filename}"
        
        log.info(f"成功生成Excel文件: {output_path}，包含 {len(test_cases)} 个测试用例")
        
        return f"""Excel文件已成功生成！

📊 **文件信息：**
- 文件名：{filename}
- 包含测试用例数量：{len(test_cases)}
- 文件路径：{output_path}

🔗 **下载链接：**
[点击下载Excel文件]({download_url})

💡 提示：如果链接无法点击，请复制以下URL到浏览器中打开：
{download_url}"""
    
    except Exception as e:
        error_msg = f"生成Excel文件失败: {str(e)}"
        log.error(error_msg)
        return error_msg

